"""
첨부 엑셀(초코펫하우스 데일리보고서)을 Supabase 시드 SQL로 변환한다.

사용법:
    python scripts/build_seed.py <엑셀경로> [출력경로]

기본 출력: supabase/seed/0001_seed_from_excel.sql
"""

import sys
import re
from datetime import datetime
from openpyxl import load_workbook

EXCLUDED_LABELS = {"파워링크", "브랜드 검색", "브랜드검색", "쇼핑브랜드형", "쇼핑 브랜드형"}


def q(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def num(v):
    if v is None or v == "":
        return "0"
    try:
        return str(float(v)) if isinstance(v, float) else str(int(v))
    except (TypeError, ValueError):
        return "0"


def is_numeric_id(v):
    return v is not None and re.fullmatch(r"\d+", str(v).strip()) is not None


def main(src, out):
    wb = load_workbook(src, data_only=True)

    # --- 상품 id 시트 -------------------------------------------------------
    # A: 상품 ID(대부분 쇼핑몰 상품ID) / B: 기본 상품명
    # D: 소재 ID / E: 노출용 상품명 / F: 쇼핑몰 상품ID
    ws = wb["상품 id"]
    base_names = {}       # product_id -> 기본 상품명
    creatives = []        # (소재ID, 노출상품명, 쇼핑몰상품ID)
    display_names = {}    # 쇼핑몰상품ID -> 노출용 상품명
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is not None and str(r[0]).strip() != "":
            base_names[str(r[0]).strip()] = (r[1] or "").strip()
        if r[3] is not None and str(r[3]).strip() != "":
            cid = str(r[3]).strip()
            mall = str(r[5]).strip() if r[5] is not None else None
            creatives.append((cid, (r[4] or "").strip(), mall))
            if mall and mall not in display_names:
                display_names[mall] = (r[4] or "").strip()

    # --- 상품군 분류 시트 ---------------------------------------------------
    # B: 상품명(기본 상품명) / C: 그룹명 / E: 상품군 종류 마스터
    ws = wb["상품군 분류"]
    group_master = []
    name_to_group = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] and r[2]:
            name_to_group[str(r[1]).strip()] = str(r[2]).strip()
        if r[4]:
            group_master.append(str(r[4]).strip())
    for g in name_to_group.values():
        if g not in group_master:
            group_master.append(g)
    group_master = list(dict.fromkeys(group_master))

    # --- result2 시트 -------------------------------------------------------
    ws = wb["result2"]
    perf = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or str(r[0]).strip() == "":
            continue
        creative_id = str(r[0]).strip()
        stat_date = r[10]
        if isinstance(stat_date, datetime):
            stat_date = stat_date.date().isoformat()
        elif stat_date:
            stat_date = str(stat_date)[:10]
        else:
            continue
        perf.append(
            {
                "creative_id": creative_id,
                "product_ref": str(r[1]).strip() if r[1] is not None else None,
                "impressions": r[2],
                "clicks": r[3],
                "cost": r[4],
                "avg_rank": r[5],
                "conv_count": r[6],
                "conv_revenue": r[7],
                "total_conv_count": r[8],
                "total_conv_revenue": r[9],
                "stat_date": stat_date,
            }
        )

    # --- 중복 제거 ----------------------------------------------------------
    # DB 의 dedupe_key 와 같은 기준으로 미리 합쳐 둔다.
    # (같은 INSERT 안에 중복 키가 있으면 ON CONFLICT DO UPDATE 가 실패한다)
    merged = {}
    for p in perf:
        ref = p["product_ref"]
        dk = ref if (is_numeric_id(ref)) else (ref if ref and ref != "0" else "-")
        key = (p["stat_date"], p["creative_id"], dk)
        if key not in merged:
            merged[key] = p
            continue
        prev = merged[key]
        total_imp = (prev["impressions"] or 0) + (p["impressions"] or 0)
        if total_imp:
            prev["avg_rank"] = (
                (prev["avg_rank"] or 0) * (prev["impressions"] or 0)
                + (p["avg_rank"] or 0) * (p["impressions"] or 0)
            ) / total_imp
        for f in ("impressions", "clicks", "cost", "conv_count", "conv_revenue",
                  "total_conv_count", "total_conv_revenue"):
            prev[f] = (prev[f] or 0) + (p[f] or 0)
    duplicates = len(perf) - len(merged)
    perf = list(merged.values())

    # --- 상품 목록 구성 -----------------------------------------------------
    # 광고에 실제로 등장했거나 매핑표에 존재하는 쇼핑몰 상품ID를 상품으로 만든다.
    product_ids = set()
    for _, _, mall in creatives:
        if is_numeric_id(mall):
            product_ids.add(mall)
    for p in perf:
        if is_numeric_id(p["product_ref"]) and p["product_ref"] != "0":
            product_ids.add(p["product_ref"])

    lines = []
    w = lines.append
    w("-- 첨부 엑셀에서 자동 생성된 시드 데이터입니다. scripts/build_seed.py 로 재생성하세요.")
    w("-- 생성 시각: " + datetime.now().isoformat(timespec="seconds"))
    w("begin;")
    w("")
    w("-- 광고 계정 ------------------------------------------------------------")
    w(
        "insert into public.ad_accounts (id, name, platform, customer_id, is_active) values "
        "('11111111-1111-1111-1111-111111111111', '초코펫하우스', 'naver_searchad', 'DEMO-CUSTOMER', true) "
        "on conflict (id) do nothing;"
    )
    w("")

    w("-- 상품군 ---------------------------------------------------------------")
    vals = ",\n  ".join(f"({q(g)}, true)" for g in group_master)
    w(f"insert into public.product_groups (name, is_active) values\n  {vals}\non conflict (name) do nothing;")
    w("")

    w("-- 상품명 -> 상품군 분류 규칙 -------------------------------------------")
    rows = []
    for base, grp in name_to_group.items():
        rows.append(f"({q(base)}, {q(grp)})")
    chunk = 200
    for i in range(0, len(rows), chunk):
        vals = ",\n  ".join(rows[i : i + chunk])
        w(
            "insert into public.product_group_rules (base_name, group_id)\n"
            f"select v.base_name, g.id from (values\n  {vals}\n) as v(base_name, group_name)\n"
            "join public.product_groups g on g.name = v.group_name\n"
            "on conflict (base_name) do nothing;"
        )
        w("")

    w("-- 상품 ---------------------------------------------------------------")
    rows = []
    for pid in sorted(product_ids):
        base = base_names.get(pid) or display_names.get(pid) or ""
        disp = display_names.get(pid) or base
        rows.append(f"({q(pid)}, {q(disp)}, {q(base)})")
    for i in range(0, len(rows), chunk):
        vals = ",\n  ".join(rows[i : i + chunk])
        w(
            "insert into public.products (mall_product_id, display_name, base_name, product_group_id)\n"
            f"select v.mall_product_id, v.display_name, v.base_name, r.group_id from (values\n  {vals}\n"
            ") as v(mall_product_id, display_name, base_name)\n"
            "left join public.product_group_rules r on r.base_name = v.base_name\n"
            "on conflict (mall_product_id) do nothing;"
        )
        w("")

    w("-- 소재 및 소재-상품 매핑 ----------------------------------------------")
    seen = set()
    crows = []
    mrows = []
    for cid, disp, mall in creatives:
        if cid in seen:
            continue
        seen.add(cid)
        crows.append(f"({q(cid)}, {q(disp)})")
        if is_numeric_id(mall):
            mrows.append(f"({q(cid)}, {q(mall)})")
    for i in range(0, len(crows), chunk):
        vals = ",\n  ".join(crows[i : i + chunk])
        w(
            "insert into public.creatives (ad_account_id, creative_id, creative_name) "
            "select '11111111-1111-1111-1111-111111111111', v.creative_id, v.creative_name from (values\n"
            f"  {vals}\n) as v(creative_id, creative_name)\n"
            "on conflict (ad_account_id, creative_id) do nothing;"
        )
        w("")
    for i in range(0, len(mrows), chunk):
        vals = ",\n  ".join(mrows[i : i + chunk])
        w(
            "insert into public.creative_product_mappings (ad_account_id, creative_id, product_id, is_active)\n"
            "select '11111111-1111-1111-1111-111111111111', v.creative_id, p.id, true from (values\n"
            f"  {vals}\n) as v(creative_id, mall_product_id)\n"
            "join public.products p on p.mall_product_id = v.mall_product_id\n"
            "on conflict (ad_account_id, creative_id) do nothing;"
        )
        w("")

    w("-- 일별 RAW 성과 데이터 -------------------------------------------------")
    prows = []
    for p in perf:
        ref = p["product_ref"]
        ad_type = "NULL"
        if ref is not None and not is_numeric_id(ref):
            ad_type = q(ref)
        elif ref == "0" or ref is None:
            ad_type = "NULL"
        mall = q(ref) if is_numeric_id(ref) and ref != "0" else "NULL"
        prows.append(
            "("
            + ", ".join(
                [
                    q(p["stat_date"]),
                    q(p["creative_id"]),
                    mall,
                    ad_type,
                    num(p["impressions"]),
                    num(p["clicks"]),
                    num(p["cost"]),
                    num(p["avg_rank"]),
                    num(p["conv_count"]),
                    num(p["conv_revenue"]),
                    num(p["total_conv_count"]),
                    num(p["total_conv_revenue"]),
                ]
            )
            + ")"
        )
    for i in range(0, len(prows), chunk):
        vals = ",\n  ".join(prows[i : i + chunk])
        w(
            "insert into public.ad_performance_daily (\n"
            "  ad_account_id, stat_date, creative_id, product_id, mall_product_id, ad_type_label,\n"
            "  impressions, clicks, cost, avg_rank, conv_count, conv_revenue,\n"
            "  total_conv_count, total_conv_revenue, source)\n"
            "select '11111111-1111-1111-1111-111111111111', v.stat_date::date, v.creative_id,\n"
            "       coalesce(p.id, pm.id), v.mall_product_id, v.ad_type_label,\n"
            "       v.impressions, v.clicks, v.cost, v.avg_rank, v.conv_count, v.conv_revenue,\n"
            "       v.total_conv_count, v.total_conv_revenue, 'excel_import'\n"
            "from (values\n"
            f"  {vals}\n"
            ") as v(stat_date, creative_id, mall_product_id, ad_type_label, impressions, clicks, cost,\n"
            "       avg_rank, conv_count, conv_revenue, total_conv_count, total_conv_revenue)\n"
            "left join public.products p on p.mall_product_id = v.mall_product_id\n"
            "left join public.creative_product_mappings m\n"
            "  on m.creative_id = v.creative_id and m.ad_account_id = '11111111-1111-1111-1111-111111111111'\n"
            "left join public.products pm on pm.id = m.product_id\n"
            "on conflict (ad_account_id, stat_date, creative_id, dedupe_key) do update set\n"
            "  impressions = excluded.impressions, clicks = excluded.clicks, cost = excluded.cost,\n"
            "  avg_rank = excluded.avg_rank, conv_count = excluded.conv_count,\n"
            "  conv_revenue = excluded.conv_revenue, total_conv_count = excluded.total_conv_count,\n"
            "  total_conv_revenue = excluded.total_conv_revenue, updated_at = now();"
        )
        w("")

    w("commit;")

    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"wrote {out}")
    print(f"  상품군 {len(group_master)} / 분류규칙 {len(name_to_group)} / 상품 {len(product_ids)}")
    print(f"  소재 {len(crows)} / 소재매핑 {len(mrows)} / RAW {len(prows)} (중복 합산 {duplicates}건)")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "초코펫하우스_데일리보고서_260724.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "supabase/seed/0001_seed_from_excel.sql"
    main(src, out)
